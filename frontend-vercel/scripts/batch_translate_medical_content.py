#!/usr/bin/env python3
"""Translate Medora medical-content workbooks with the OpenAI Batch API.

The script is intentionally split into resumable stages:

    prepare -> submit -> status -> collect

The convenience ``run`` command performs every stage and waits for the Batch
job to reach a terminal state. Production database writes are deliberately out
of scope: collection only creates translated workbook copies.

Examples:

    python scripts/batch_translate_medical_content.py prepare \
      --workbook-dir /path/to/medical-content-i18n \
      --run-dir /path/to/run \
      --sheet Procedures \
      --languages ru ar id \
      --slug 25g-27g-pars-plana-vitrectomy-ppv

    python scripts/batch_translate_medical_content.py submit \
      --run-dir /path/to/run

    python scripts/batch_translate_medical_content.py status \
      --run-dir /path/to/run --wait

    python scripts/batch_translate_medical_content.py collect \
      --run-dir /path/to/run

Requirements:

    pip install -r scripts/requirements-batch-translation.txt
    export OPENAI_API_KEY=...
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence


DEFAULT_MODEL = "gpt-5.6-terra"
DEFAULT_ENDPOINT = "/v1/chat/completions"
DEFAULT_LANGUAGES = ("ru", "ar", "id")
DEFAULT_SHEET = "Procedures"
SUPPORTED_SOURCE_LOCALES = {"en", "zh"}
SOURCE_LANGUAGE_NAMES = {
    "en": "English",
    "zh": "Simplified Chinese",
}
HEADER_ROW = 3
DATA_START_ROW = 4
TERMINAL_BATCH_STATUSES = {"completed", "failed", "expired", "cancelled"}
RUNNING_BATCH_STATUSES = {"validating", "in_progress", "finalizing", "cancelling"}

LANGUAGES: dict[str, dict[str, str]] = {
    "en": {
        "english_name": "English",
        "native_name": "English",
        "style": (
            "Use clear international English written for patients. Use established "
            "clinical terminology, expand uncommon abbreviations on first use when "
            "the source provides the expanded term, and avoid marketing language."
        ),
    },
    "es": {
        "english_name": "Spanish",
        "native_name": "Español",
        "style": (
            "Use neutral international Spanish written for patients. Use established "
            "medical terminology understood across Spanish-speaking regions and avoid "
            "unnecessary regionalisms."
        ),
    },
    "fr": {
        "english_name": "French",
        "native_name": "Français",
        "style": (
            "Use clear international French written for patients. Use established "
            "French clinical terminology and natural sentence structure rather than "
            "literal source-language phrasing."
        ),
    },
    "de": {
        "english_name": "German",
        "native_name": "Deutsch",
        "style": (
            "Use clear standard German written for patients. Use established German "
            "clinical terminology, natural compound nouns, and a formal but accessible "
            "tone."
        ),
    },
    "ru": {
        "english_name": "Russian",
        "native_name": "Русский",
        "style": (
            "Use natural, formal Russian written for patients. Use established "
            "Russian clinical terminology and avoid literal English calques. "
            "For example, translate 'implantable cardioverter-defibrillator (ICD)' "
            "as 'имплантируемый кардиовертер-дефибриллятор (ICD)': keep the Latin "
            "acronym unchanged. Translate 'en bloc resection of a spinal tumor' as "
            "'резекция опухоли позвоночника единым блоком'."
        ),
    },
    "ar": {
        "english_name": "Modern Standard Arabic",
        "native_name": "العربية الفصحى",
        "style": (
            "Use clear Modern Standard Arabic written for international patients. "
            "Use established Arabic medical terminology; retain a standard English "
            "abbreviation in parentheses when it improves clinical clarity. Use "
            "'شريط منتصف الإحليل (MUS)' for 'mid-urethral sling (MUS)' and "
            "'ورم العمود الفقري' for 'spinal tumor' unless the source explicitly "
            "refers to the spinal cord. Translate 'posterior thoracic decompression' "
            "as 'إزالة الضغط الخلفية عن العمود الفقري الصدري'."
        ),
    },
    "id": {
        "english_name": "Indonesian",
        "native_name": "Bahasa Indonesia",
        "style": (
            "Use natural, formal Indonesian written for patients. Use terminology "
            "commonly used in Indonesian clinical and hospital communication. "
            "Translate the expanded clinical term rather than leaving it in English; "
            "for example, use 'kardioverter-defibrilator implan (ICD)' while keeping "
            "the Latin acronym unchanged, and use 'sling miduretra (MUS)' for "
            "'mid-urethral sling (MUS)'."
        ),
    },
}

TRANSLATABLE_FIELDS = (
    "name",
    "waiting_time",
    "cost_coverage",
    "cost_factors",
    "stay_at_hospital",
    "stay_at_hotel",
    "stay_in_china",
    "surgery_detailed_description",
    "when_is_needed",
    "preparation_before_surgery",
    "recovery_process",
    "surgery_options",
    "faqs_json",
    "surgery_steps_json",
    "recovery_steps_json",
)

COPY_FIELDS = (
    "cost_usd",
    "cost_as_of",
    "cost_estimate",
)

JSON_FIELDS = {
    "faqs_json",
    "surgery_steps_json",
    "recovery_steps_json",
}

# A small set of identity-bearing abbreviations is enforced in procedure names.
# Generic acronym preservation produced false positives for clinically correct
# localized forms (for example ASD -> ДМПП and TURBT -> ТУР in Russian), as well
# as tokens such as CO extracted from CO₂. The prompt still asks the model to
# retain useful Latin abbreviations; this list is only the hard safety gate.
MANDATORY_NAME_ACRONYMS = {
    "CABG",
    "ICD",
    "MUS",
    "PCI",
    "PPV",
}


class TranslationError(RuntimeError):
    """Raised when a translation run cannot safely continue."""


@dataclass(frozen=True)
class WorkbookRow:
    workbook_path: Path
    sheet_name: str
    row_number: int
    entity_id: str
    slug: str
    source_locale: str
    target_locale: str
    source: dict[str, Any]
    copied: dict[str, Any]
    locked_targets: dict[str, Any]
    source_hash: str


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha256_json(value: Any) -> str:
    payload = json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def atomic_write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(value, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def object_to_dict(value: Any) -> dict[str, Any]:
    if hasattr(value, "model_dump"):
        return value.model_dump(mode="json")
    if isinstance(value, dict):
        return value
    raise TranslationError(f"Cannot serialize API object of type {type(value).__name__}")


def require_openpyxl() -> tuple[Any, Any]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise TranslationError(
            "openpyxl is required. Install scripts/requirements-batch-translation.txt"
        ) from exc
    return load_workbook, get_column_letter


def make_openai_client() -> Any:
    if not os.getenv("OPENAI_API_KEY"):
        raise TranslationError("OPENAI_API_KEY is not set")
    try:
        from openai import OpenAI
    except ImportError as exc:
        raise TranslationError(
            "The OpenAI Python SDK is required. Install "
            "scripts/requirements-batch-translation.txt"
        ) from exc

    # Batch execution itself is managed by OpenAI. Disabling SDK request retries
    # keeps this CLI's behavior explicit and prevents a hidden retry loop.
    return OpenAI(max_retries=0, timeout=60.0)


def api_call(description: str, function: Any, *args: Any, **kwargs: Any) -> Any:
    """Execute one SDK operation without retries and normalize transport errors."""
    try:
        return function(*args, **kwargs)
    except Exception as exc:
        raise TranslationError(f"{description} failed: {exc}") from exc


def normalize_languages(values: Sequence[str]) -> list[str]:
    normalized: list[str] = []
    for raw in values:
        for code in raw.split(","):
            code = code.strip().lower()
            if not code:
                continue
            if code not in LANGUAGES:
                raise TranslationError(
                    f"Unsupported language {code!r}; choose from {sorted(LANGUAGES)}"
                )
            if code not in normalized:
                normalized.append(code)
    if not normalized:
        raise TranslationError("At least one target language is required")
    return normalized


def resolve_workbook(workbook_dir: Path, locale: str) -> Path:
    expected = workbook_dir / f"medora_medical_content_{locale}.xlsx"
    if expected.is_file():
        return expected.resolve()

    candidates = sorted(workbook_dir.glob(f"*_{locale}.xlsx"))
    if len(candidates) == 1:
        return candidates[0].resolve()
    if not candidates:
        raise TranslationError(
            f"No workbook found for {locale!r} in {workbook_dir}"
        )
    raise TranslationError(
        f"Multiple workbooks found for {locale!r}: "
        + ", ".join(str(path) for path in candidates)
    )


def header_map(worksheet: Any) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column in range(1, worksheet.max_column + 1):
        value = worksheet.cell(HEADER_ROW, column).value
        if isinstance(value, str) and value.strip():
            headers[value.strip()] = column
    return headers


def validate_headers(headers: Mapping[str, int]) -> None:
    required = {
        "entity_id",
        "slug",
        "source_locale",
        "target_locale",
        "status",
        "reviewer_notes",
    }
    for field in (*TRANSLATABLE_FIELDS, *COPY_FIELDS):
        required.add(f"source_{field}")
        required.add(f"target_{field}")
    missing = sorted(required - set(headers))
    if missing:
        raise TranslationError(
            "Workbook is missing required columns: " + ", ".join(missing)
        )


def parse_source_field(field: str, value: Any, context: str) -> Any:
    if value is None or value == "":
        return None
    if field not in JSON_FIELDS:
        return str(value)
    if not isinstance(value, str):
        raise TranslationError(f"{context}: {field} must contain JSON text")
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError as exc:
        raise TranslationError(f"{context}: invalid {field}: {exc}") from exc
    if not isinstance(parsed, (dict, list)):
        raise TranslationError(f"{context}: {field} must be a JSON object or array")
    if not parsed:
        raise TranslationError(f"{context}: {field} must not be empty")
    return parsed


def normalize_copied_field(field: str, value: Any) -> Any:
    if field != "cost_usd" or not isinstance(value, str):
        return value
    match = re.fullmatch(
        r"\s*约\s*¥\s*([\d,]+(?:\.\d+)?)\s*（估算）\s*",
        value,
    )
    if not match:
        return value
    return f"CNY {match.group(1)}"


def parse_locked_target_fields(value: Any, context: str) -> set[str]:
    if value is None or str(value).strip() == "":
        return set()
    fields = {
        field.strip()
        for field in str(value).split(",")
        if field.strip()
    }
    unsupported = sorted(fields - set(TRANSLATABLE_FIELDS))
    if unsupported:
        raise TranslationError(
            f"{context}: locked_target_fields contains unsupported fields: "
            + ", ".join(unsupported)
        )
    return fields


def row_matches(
    worksheet: Any,
    headers: Mapping[str, int],
    row_number: int,
    slugs: set[str],
    include_complete: bool,
) -> bool:
    slug = str(worksheet.cell(row_number, headers["slug"]).value or "").strip()
    if not slug:
        return False
    if slugs and slug not in slugs:
        return False
    status = str(
        worksheet.cell(row_number, headers["status"]).value or ""
    ).strip().lower()
    if status == "complete" and not include_complete:
        return False
    return True


def extract_rows(
    workbook_path: Path,
    sheet_name: str,
    slugs: set[str],
    limit: int | None,
    include_complete: bool,
) -> list[WorkbookRow]:
    load_workbook, _ = require_openpyxl()
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    if sheet_name not in workbook.sheetnames:
        raise TranslationError(f"{workbook_path}: missing sheet {sheet_name!r}")

    worksheet = workbook[sheet_name]
    headers = header_map(worksheet)
    validate_headers(headers)
    rows: list[WorkbookRow] = []

    for row_number in range(DATA_START_ROW, worksheet.max_row + 1):
        if not row_matches(
            worksheet,
            headers,
            row_number,
            slugs,
            include_complete,
        ):
            continue

        entity_id = str(
            worksheet.cell(row_number, headers["entity_id"]).value or ""
        ).strip()
        slug = str(worksheet.cell(row_number, headers["slug"]).value or "").strip()
        source_locale = str(
            worksheet.cell(row_number, headers["source_locale"]).value or ""
        ).strip()
        target_locale = str(
            worksheet.cell(row_number, headers["target_locale"]).value or ""
        ).strip()
        context = f"{workbook_path.name}:{sheet_name}!row {row_number} ({slug})"

        if not entity_id:
            raise TranslationError(f"{context}: entity_id is empty")
        if source_locale not in SUPPORTED_SOURCE_LOCALES:
            raise TranslationError(
                f"{context}: unsupported source_locale {source_locale!r}; "
                f"choose from {sorted(SUPPORTED_SOURCE_LOCALES)}"
            )
        if target_locale not in LANGUAGES:
            raise TranslationError(
                f"{context}: unsupported target_locale {target_locale!r}"
            )

        source: dict[str, Any] = {}
        copied: dict[str, Any] = {}
        locked_targets: dict[str, Any] = {}
        for field in TRANSLATABLE_FIELDS:
            value = worksheet.cell(
                row_number, headers[f"source_{field}"]
            ).value
            parsed = parse_source_field(field, value, context)
            if parsed is not None:
                source[field] = parsed
        locked_fields = parse_locked_target_fields(
            worksheet.cell(
                row_number,
                headers["locked_target_fields"],
            ).value
            if "locked_target_fields" in headers
            else None,
            context,
        )
        for field in locked_fields:
            value = worksheet.cell(
                row_number,
                headers[f"target_{field}"],
            ).value
            parsed = parse_source_field(field, value, context)
            if parsed is None:
                raise TranslationError(
                    f"{context}: target_{field} is blank but the field is locked"
                )
            locked_targets[field] = parsed
        for field in COPY_FIELDS:
            value = worksheet.cell(
                row_number, headers[f"source_{field}"]
            ).value
            if value is not None:
                copied[field] = normalize_copied_field(field, value)

        missing_source_fields = [
            field for field in TRANSLATABLE_FIELDS if field not in source
        ]
        if missing_source_fields:
            raise TranslationError(
                f"{context}: incomplete procedure source; missing "
                + ", ".join(missing_source_fields)
            )

        source_identity = {
            "entity_id": entity_id,
            "slug": slug,
            "source_locale": source_locale,
            "target_locale": target_locale,
            "source": source,
            "copied": copied,
            "locked_targets": locked_targets,
        }
        rows.append(
            WorkbookRow(
                workbook_path=workbook_path.resolve(),
                sheet_name=sheet_name,
                row_number=row_number,
                entity_id=entity_id,
                slug=slug,
                source_locale=source_locale,
                target_locale=target_locale,
                source=source,
                copied=copied,
                locked_targets=locked_targets,
                source_hash=sha256_json(source_identity),
            )
        )
        if limit is not None and len(rows) >= limit:
            break

    workbook.close()
    return rows


def build_system_prompt(locale: str, source_locale: str = "en") -> str:
    language = LANGUAGES[locale]
    source_language = SOURCE_LANGUAGE_NAMES[source_locale]
    return f"""You are a senior medical translator for an international patient platform.

Translate the supplied {source_language} medical procedure content into {language["english_name"]} ({language["native_name"]}).

Rules:
1. Return exactly one valid JSON object and no surrounding prose or Markdown.
2. Preserve the top-level field set exactly: do not add, remove, rename, or reorder content conceptually.
3. Translate every populated human-readable value. Do not summarize, omit, expand, fact-check, or invent medical information.
4. Preserve numbers, ranges, units, currency amounts, dates, and HTML. In the procedure name, preserve every Latin-script medical acronym exactly. In body copy, preserve procedure, device, therapy, and biomarker acronyms (for example ICD, MUS, PPV, TRUS, HLA, and VEGF). Common diagnostic modality acronyms such as CT, MRI, ECG, EEG, PET, SPECT, US, or USG may use the standard localized form. Translate the surrounding expanded term. Preserve hospital, university, organization, product, and brand entity names unless a standard target-language form exists.
5. For faqs_json, the object keys are patient-facing questions: translate both each question key and its answer value while preserving the number of FAQ entries.
6. For surgery_steps_json and recovery_steps_json, preserve structural keys such as "steps", "step", "text", and "guidance", preserve step numbers and list length, and translate human-readable "text" and "guidance" values.
7. Never return the source language as a fallback. If a phrase is normally retained in the source language, keep it only where professionally appropriate.
8. Maintain a neutral, medically responsible tone. Do not introduce treatment promises, success rates, comparative claims, or personalized medical advice.
9. {language["style"]}
"""


def build_request_body(row: WorkbookRow, model: str) -> dict[str, Any]:
    return {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": build_system_prompt(
                    row.target_locale,
                    row.source_locale,
                ),
            },
            {
                "role": "user",
                "content": json.dumps(
                    row.source,
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
            },
        ],
        "response_format": {"type": "json_object"},
        "reasoning_effort": "none",
        "max_completion_tokens": 20000,
    }


def custom_id_for(row: WorkbookRow) -> str:
    slug_token = re.sub(r"[^a-zA-Z0-9_-]+", "-", row.slug).strip("-")
    slug_token = slug_token[:64] or "procedure"
    digest = hashlib.sha256(
        f"{row.entity_id}|{row.target_locale}|{row.source_hash}".encode("utf-8")
    ).hexdigest()[:12]
    return f"procedure--{row.target_locale}--{slug_token}--{digest}"


def prepare_run(args: argparse.Namespace) -> dict[str, Any]:
    workbook_dir = Path(args.workbook_dir).expanduser().resolve()
    run_dir = Path(args.run_dir).expanduser().resolve()
    manifest_path = run_dir / "manifest.json"
    input_path = run_dir / "batch_input.jsonl"

    if manifest_path.exists():
        existing_manifest = read_json(manifest_path)
        if existing_manifest.get("batch_id"):
            raise TranslationError(
                f"{manifest_path} belongs to submitted batch "
                f"{existing_manifest['batch_id']}; use a different run-dir"
            )
        if not args.force:
            raise TranslationError(
                f"{manifest_path} already exists; resume it or use --force "
                "to replace the unsubmitted run"
            )

    languages = normalize_languages(args.languages)
    slugs = {value.strip() for value in args.slug if value.strip()}
    if args.limit is not None and args.limit < 1:
        raise TranslationError("--limit must be at least 1")

    run_dir.mkdir(parents=True, exist_ok=True)
    rows: list[WorkbookRow] = []
    workbook_paths: dict[str, str] = {}
    for locale in languages:
        workbook_path = resolve_workbook(workbook_dir, locale)
        workbook_paths[locale] = str(workbook_path)
        locale_rows = extract_rows(
            workbook_path=workbook_path,
            sheet_name=args.sheet,
            slugs=slugs,
            limit=args.limit,
            include_complete=args.include_complete,
        )
        if not locale_rows:
            raise TranslationError(
                f"No matching rows found in {workbook_path.name}; "
                "check the slug/status filters"
            )
        wrong_locale = [
            row for row in locale_rows if row.target_locale != locale
        ]
        if wrong_locale:
            raise TranslationError(
                f"{workbook_path.name} contains rows for unexpected target locales"
            )
        rows.extend(locale_rows)

    requests: list[dict[str, Any]] = []
    entries: dict[str, dict[str, Any]] = {}
    for row in rows:
        custom_id = custom_id_for(row)
        if custom_id in entries:
            raise TranslationError(f"Duplicate custom_id generated: {custom_id}")
        request = {
            "custom_id": custom_id,
            "method": "POST",
            "url": DEFAULT_ENDPOINT,
            "body": build_request_body(row, args.model),
        }
        requests.append(request)
        entries[custom_id] = {
            "workbook_path": str(row.workbook_path),
            "sheet_name": row.sheet_name,
            "row_number": row.row_number,
            "entity_id": row.entity_id,
            "slug": row.slug,
            "source_locale": row.source_locale,
            "target_locale": row.target_locale,
            "source": row.source,
            "copied": row.copied,
            "locked_targets": row.locked_targets,
            "source_hash": row.source_hash,
            "expected_fields": list(row.source.keys()),
        }

    with input_path.open("w", encoding="utf-8") as handle:
        for request in requests:
            handle.write(
                json.dumps(request, ensure_ascii=False, separators=(",", ":"))
                + "\n"
            )

    run_id = str(uuid.uuid4())
    manifest = {
        "schema_version": 1,
        "run_id": run_id,
        "created_at": utc_now(),
        "updated_at": utc_now(),
        "status": "prepared",
        "model": args.model,
        "endpoint": DEFAULT_ENDPOINT,
        "completion_window": "24h",
        "workbook_dir": str(workbook_dir),
        "workbooks": workbook_paths,
        "sheet": args.sheet,
        "languages": languages,
        "filters": {
            "slugs": sorted(slugs),
            "limit_per_language": args.limit,
            "include_complete": args.include_complete,
        },
        "request_count": len(requests),
        "input_file": input_path.name,
        "input_sha256": hashlib.sha256(input_path.read_bytes()).hexdigest(),
        "input_file_id": None,
        "batch_id": None,
        "batch_status": None,
        "output_file_id": None,
        "error_file_id": None,
        "entries": entries,
    }
    atomic_write_json(manifest_path, manifest)
    print(
        f"Prepared {len(requests)} requests for "
        f"{len(languages)} language(s): {input_path}"
    )
    return manifest


def load_manifest(run_dir: Path) -> tuple[Path, dict[str, Any]]:
    manifest_path = run_dir / "manifest.json"
    if not manifest_path.is_file():
        raise TranslationError(f"Manifest not found: {manifest_path}")
    manifest = read_json(manifest_path)
    if manifest.get("schema_version") != 1:
        raise TranslationError("Unsupported manifest schema version")
    return manifest_path, manifest


def save_manifest(path: Path, manifest: dict[str, Any]) -> None:
    manifest["updated_at"] = utc_now()
    atomic_write_json(path, manifest)


def find_existing_batch(client: Any, run_id: str) -> Any | None:
    page = api_call("listing recent batches", client.batches.list, limit=100)
    for batch in getattr(page, "data", []):
        metadata = getattr(batch, "metadata", None) or {}
        if metadata.get("medora_run_id") == run_id:
            return batch
    return None


def submit_run(run_dir: Path) -> dict[str, Any]:
    manifest_path, manifest = load_manifest(run_dir)
    client = make_openai_client()

    if manifest.get("batch_id"):
        batch = api_call(
            "retrieving submitted batch",
            client.batches.retrieve,
            manifest["batch_id"],
        )
        manifest["batch_status"] = batch.status
        manifest["status"] = "submitted"
        save_manifest(manifest_path, manifest)
        print(f"Batch already submitted: {batch.id} ({batch.status})")
        return manifest

    input_path = run_dir / manifest["input_file"]
    expected_hash = manifest["input_sha256"]
    actual_hash = hashlib.sha256(input_path.read_bytes()).hexdigest()
    if actual_hash != expected_hash:
        raise TranslationError(
            "Batch input changed after preparation; refusing to submit"
        )

    if not manifest.get("input_file_id"):
        with input_path.open("rb") as handle:
            uploaded = api_call(
                "uploading Batch input file",
                client.files.create,
                file=handle,
                purpose="batch",
            )
        manifest["input_file_id"] = uploaded.id
        manifest["status"] = "uploaded"
        save_manifest(manifest_path, manifest)
        print(f"Uploaded batch input: {uploaded.id}")

    existing = find_existing_batch(client, manifest["run_id"])
    if existing is not None:
        batch = existing
        print(f"Recovered existing batch for run_id: {batch.id}")
    else:
        batch = api_call(
            "creating Batch job",
            client.batches.create,
            input_file_id=manifest["input_file_id"],
            endpoint=manifest["endpoint"],
            completion_window=manifest["completion_window"],
            metadata={
                "description": "Medora medical content translation",
                "medora_run_id": manifest["run_id"],
                "model": manifest["model"],
                "languages": ",".join(manifest["languages"]),
            },
        )

    manifest["batch_id"] = batch.id
    manifest["batch_status"] = batch.status
    manifest["status"] = "submitted"
    save_manifest(manifest_path, manifest)
    atomic_write_json(run_dir / "batch_created.json", object_to_dict(batch))
    print(f"Submitted batch: {batch.id} ({batch.status})")
    return manifest


def refresh_batch(run_dir: Path) -> tuple[dict[str, Any], Any]:
    manifest_path, manifest = load_manifest(run_dir)
    batch_id = manifest.get("batch_id")
    if not batch_id:
        raise TranslationError("Run has not been submitted")
    client = make_openai_client()
    batch = api_call(
        "retrieving Batch status",
        client.batches.retrieve,
        batch_id,
    )
    manifest["batch_status"] = batch.status
    manifest["output_file_id"] = getattr(batch, "output_file_id", None)
    manifest["error_file_id"] = getattr(batch, "error_file_id", None)
    manifest["status"] = (
        "batch_terminal"
        if batch.status in TERMINAL_BATCH_STATUSES
        else "submitted"
    )
    save_manifest(manifest_path, manifest)
    atomic_write_json(run_dir / "batch_status.json", object_to_dict(batch))
    return manifest, batch


def status_run(
    run_dir: Path,
    wait: bool,
    poll_interval: int,
) -> dict[str, Any]:
    if poll_interval < 5:
        raise TranslationError("--poll-interval must be at least 5 seconds")

    while True:
        manifest, batch = refresh_batch(run_dir)
        counts = getattr(batch, "request_counts", None)
        if counts:
            progress = (
                f"total={getattr(counts, 'total', '?')} "
                f"completed={getattr(counts, 'completed', '?')} "
                f"failed={getattr(counts, 'failed', '?')}"
            )
        else:
            progress = "request counts unavailable"
        print(f"Batch {batch.id}: {batch.status}; {progress}", flush=True)

        if batch.status in TERMINAL_BATCH_STATUSES:
            return manifest
        if batch.status not in RUNNING_BATCH_STATUSES:
            raise TranslationError(f"Unexpected batch status: {batch.status}")
        if not wait:
            return manifest
        time.sleep(poll_interval)


def download_file(client: Any, file_id: str, output_path: Path) -> None:
    response = api_call(
        f"downloading API file {file_id}",
        client.files.content,
        file_id,
    )
    content = getattr(response, "content", None)
    if content is None:
        text = getattr(response, "text", None)
        if text is None:
            raise TranslationError(f"Could not read downloaded file {file_id}")
        content = text.encode("utf-8")
    if isinstance(content, str):
        content = content.encode("utf-8")
    output_path.write_bytes(content)


def parse_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            try:
                value = json.loads(line)
            except json.JSONDecodeError as exc:
                raise TranslationError(
                    f"{path.name}:{line_number}: invalid JSON: {exc}"
                ) from exc
            if not isinstance(value, dict):
                raise TranslationError(
                    f"{path.name}:{line_number}: expected a JSON object"
                )
            rows.append(value)
    return rows


def extract_chat_content(result: Mapping[str, Any]) -> str:
    response = result.get("response")
    if not isinstance(response, dict):
        raise TranslationError("Batch result has no response object")
    if response.get("status_code") != 200:
        raise TranslationError(
            f"Batch request returned HTTP {response.get('status_code')}"
        )
    body = response.get("body")
    if not isinstance(body, dict):
        raise TranslationError("Batch result has no response body")
    choices = body.get("choices")
    if not isinstance(choices, list) or not choices:
        raise TranslationError("Batch response has no choices")
    message = choices[0].get("message")
    if not isinstance(message, dict):
        raise TranslationError("Batch response has no assistant message")
    refusal = message.get("refusal")
    if refusal:
        raise TranslationError(f"Model refused translation: {refusal}")
    content = message.get("content")
    if not isinstance(content, str) or not content.strip():
        raise TranslationError("Batch response content is empty")
    return content


def validate_steps_structure(
    source: Any,
    translated: Any,
    field: str,
) -> list[str]:
    errors: list[str] = []
    if not isinstance(source, dict) or not isinstance(translated, dict):
        return [f"{field}: expected source and target JSON objects"]
    if set(source) != set(translated):
        errors.append(f"{field}: root JSON keys changed")
        return errors
    source_steps = source.get("steps")
    target_steps = translated.get("steps")
    if not isinstance(source_steps, list) or not isinstance(target_steps, list):
        return [f"{field}: steps must remain arrays"]
    if len(source_steps) != len(target_steps):
        errors.append(
            f"{field}: step count changed "
            f"from {len(source_steps)} to {len(target_steps)}"
        )
        return errors
    for index, (source_step, target_step) in enumerate(
        zip(source_steps, target_steps),
        start=1,
    ):
        if not isinstance(source_step, dict) or not isinstance(target_step, dict):
            errors.append(f"{field}: step {index} must remain an object")
            continue
        if set(source_step) != set(target_step):
            errors.append(f"{field}: step {index} keys changed")
        source_number = source_step.get("step")
        target_number = target_step.get("step")
        source_is_number = isinstance(source_number, (int, float)) or (
            isinstance(source_number, str)
            and bool(re.fullmatch(r"\s*\d+(?:\.\d+)?\s*", source_number))
        )
        if source_is_number:
            try:
                numbers_match = float(source_number) == float(target_number)
            except (TypeError, ValueError):
                numbers_match = False
            if not numbers_match:
                errors.append(f"{field}: step {index} number changed")
        elif source_number not in (None, "") and target_number in (None, ""):
            # Some imported source rows contain an instruction sentence in the
            # ``step`` slot instead of a number. Translating that sentence is
            # correct; only require the target slot to remain populated.
            errors.append(f"{field}: step {index} translated step is empty")
        for key, value in source_step.items():
            if key == "step" or value in (None, ""):
                continue
            target_value = target_step.get(key)
            if target_value in (None, ""):
                errors.append(
                    f"{field}: step {index} translated {key!r} is empty"
                )
            elif (
                isinstance(value, str)
                and isinstance(target_value, str)
                and key == "text"
                and is_substantial_human_text(value)
                and (len(value.strip()) >= 25 or len(value.split()) >= 4)
                and value.strip() == target_value.strip()
            ):
                errors.append(
                    f"{field}: step {index} {key!r} is unchanged from source"
                )
    return errors


def validate_faq_structure(source: Any, translated: Any) -> list[str]:
    if not isinstance(source, dict) or not isinstance(translated, dict):
        return ["faqs_json: expected source and target JSON objects"]
    errors: list[str] = []
    if len(source) != len(translated):
        errors.append(
            f"faqs_json: FAQ count changed from {len(source)} to {len(translated)}"
        )
    for question, answer in translated.items():
        if not isinstance(question, str) or not question.strip():
            errors.append("faqs_json: translated question is empty or non-text")
        if not isinstance(answer, str) or not answer.strip():
            errors.append("faqs_json: translated answer is empty or non-text")
    if len(source) == len(translated):
        for index, ((source_question, source_answer), (question, answer)) in enumerate(
            zip(source.items(), translated.items()),
            start=1,
        ):
            if (
                is_substantial_human_text(source_question)
                and source_question.strip() == str(question).strip()
            ):
                errors.append(f"faqs_json: question {index} is unchanged from source")
            if (
                is_substantial_human_text(source_answer)
                and source_answer.strip() == str(answer).strip()
            ):
                errors.append(f"faqs_json: answer {index} is unchanged from source")
    return errors


def is_substantial_human_text(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    alphabetic = re.findall(
        r"[A-Za-zА-Яа-яЁё\u0600-\u06FF\u4E00-\u9FFF]",
        value,
    )
    return len(alphabetic) >= 8


def flatten_text(value: Any) -> Iterable[str]:
    if isinstance(value, str):
        yield value
    elif isinstance(value, dict):
        for key, child in value.items():
            if isinstance(key, str):
                yield key
            yield from flatten_text(child)
    elif isinstance(value, list):
        for child in value:
            yield from flatten_text(child)


def extract_acronyms(value: Any) -> set[str]:
    text = " ".join(flatten_text(value))
    return set(re.findall(r"(?<![A-Za-z0-9])[A-Z][A-Z0-9]{1,}(?![A-Za-z0-9])", text))


def language_signal(locale: str, value: Any) -> bool:
    text = " ".join(flatten_text(value))
    if locale == "ru":
        return bool(re.search(r"[А-Яа-яЁё]", text))
    if locale == "ar":
        return bool(re.search(r"[\u0600-\u06FF]", text))
    if locale in {"en", "es", "fr", "de", "id"}:
        # These languages share the Latin script. Substantial unchanged-source
        # checks below provide the stronger guard when the source is also Latin.
        return bool(re.search(r"[A-Za-z]", text))
    return False


def numeric_tokens(value: Any) -> list[str]:
    text = " ".join(flatten_text(value))
    digit_translation = str.maketrans(
        "٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹",
        "01234567890123456789",
    )
    normalized = text.translate(digit_translation)
    return [
        re.sub(r"[,٫،]", ".", token)
        for token in re.findall(r"\d+(?:[.,٫،]\d+)?", normalized)
    ]

def localized_number_count(locale: str, token: str, value: Any) -> int:
    if token != "1":
        return 0
    text = " ".join(flatten_text(value))
    patterns = {
        "en": r"\bone\b",
        "es": r"\b(?:un|una|uno)\b",
        "fr": r"\b(?:un|une)\b",
        "de": r"\b(?:ein|eine|einen|einem|einer|eines)\b",
        "ru": r"\b(?:один|одна|одно|одну)\b",
        "ar": r"(?:واحد(?:ة|اً)?|إحدى)",
        "id": r"\bsatu\b",
    }
    pattern = patterns.get(locale)
    if not pattern:
        return 0
    return len(re.findall(pattern, text, flags=re.IGNORECASE))


def validate_translation(
    entry: Mapping[str, Any],
    translated: Any,
) -> list[str]:
    source = entry["source"]
    locale = entry["target_locale"]
    locked_fields = set(entry.get("locked_targets", {}))
    errors: list[str] = []

    if not isinstance(translated, dict):
        return ["model output is not a JSON object"]
    if set(translated) != set(source):
        missing = sorted(set(source) - set(translated))
        extra = sorted(set(translated) - set(source))
        if missing:
            errors.append("missing fields: " + ", ".join(missing))
        if extra:
            errors.append("unexpected fields: " + ", ".join(extra))

    for field, source_value in source.items():
        if field not in translated:
            continue
        target_value = translated[field]
        if target_value in (None, ""):
            errors.append(f"{field}: translated value is empty")
            continue
        if field == "faqs_json":
            errors.extend(validate_faq_structure(source_value, target_value))
        elif field in {"surgery_steps_json", "recovery_steps_json"}:
            errors.extend(
                validate_steps_structure(source_value, target_value, field)
            )

    if not language_signal(locale, translated):
        errors.append(f"no {locale} language signal detected")

    for field, source_value in source.items():
        target_value = translated.get(field)
        if (
            field != "name"
            and is_substantial_human_text(source_value)
            and isinstance(target_value, str)
            and target_value.strip() == source_value.strip()
        ):
            errors.append(f"{field}: substantial text is unchanged from source")
        source_numbers = numeric_tokens(source_value)
        target_numbers = numeric_tokens(target_value)
        missing_numbers = sorted(
            token
            for token in set(source_numbers)
            if (
                target_numbers.count(token)
                + localized_number_count(locale, token, target_value)
                < source_numbers.count(token)
            )
        )
        if missing_numbers:
            errors.append(
                f"{field}: numeric tokens were not preserved: "
                + ", ".join(missing_numbers)
            )
        if (
            field == "name"
            and field not in locked_fields
            and target_value is not None
        ):
            target_text = " ".join(flatten_text(target_value))
            for acronym in sorted(extract_acronyms(source_value)):
                if acronym not in MANDATORY_NAME_ACRONYMS:
                    continue
                if not re.search(
                    rf"(?<![A-Za-z0-9]){re.escape(acronym)}(?![A-Za-z0-9])",
                    target_text,
                ):
                    errors.append(
                        f"{field}: source acronym {acronym!r} was not preserved"
                    )

    return errors


def merge_locked_targets(
    translated: Any,
    locked_targets: Mapping[str, Any],
) -> Any:
    if not isinstance(translated, dict):
        return translated
    return {
        **translated,
        **locked_targets,
    }


def verify_source_unchanged(
    entry: Mapping[str, Any],
    workbook_path: Path,
) -> None:
    rows = extract_rows(
        workbook_path=workbook_path,
        sheet_name=entry["sheet_name"],
        slugs={entry["slug"]},
        limit=1,
        include_complete=True,
    )
    if len(rows) != 1:
        raise TranslationError(
            f"Could not re-resolve source row for {entry['slug']}"
        )
    current = rows[0]
    if current.row_number != entry["row_number"]:
        raise TranslationError(
            f"Row moved for {entry['slug']}; prepare a new batch"
        )
    if current.source_hash != entry["source_hash"]:
        raise TranslationError(
            f"Source content changed for {entry['slug']}; prepare a new batch"
        )


def write_translations(
    run_dir: Path,
    manifest: Mapping[str, Any],
    translations: Mapping[str, Mapping[str, Any]],
) -> dict[str, str]:
    load_workbook, _ = require_openpyxl()
    output_dir = run_dir / "output_workbooks"
    output_dir.mkdir(parents=True, exist_ok=True)
    grouped: dict[Path, list[tuple[Mapping[str, Any], Mapping[str, Any]]]] = {}

    for custom_id, translated in translations.items():
        entry = manifest["entries"][custom_id]
        workbook_path = Path(entry["workbook_path"])
        grouped.setdefault(workbook_path, []).append((entry, translated))

    outputs: dict[str, str] = {}
    for source_path, records in grouped.items():
        for entry, _ in records:
            verify_source_unchanged(entry, source_path)

        workbook = load_workbook(source_path, read_only=False, data_only=False)
        for entry, translated in records:
            worksheet = workbook[entry["sheet_name"]]
            headers = header_map(worksheet)
            validate_headers(headers)
            row_number = entry["row_number"]

            for field, value in translated.items():
                target_cell = worksheet.cell(
                    row_number, headers[f"target_{field}"]
                )
                if field in JSON_FIELDS:
                    target_cell.value = json.dumps(
                        value,
                        ensure_ascii=False,
                        separators=(",", ":"),
                    )
                else:
                    target_cell.value = value

            for field, value in entry["copied"].items():
                target_cell = worksheet.cell(
                    row_number, headers[f"target_{field}"]
                )
                target_cell.value = value

            worksheet.cell(row_number, headers["status"]).value = "Complete"
            previous_note = str(
                worksheet.cell(row_number, headers["reviewer_notes"]).value or ""
            ).strip()
            audit_note = (
                f"OpenAI Batch {manifest['batch_id']} · {manifest['model']} · "
                f"collected {utc_now()}"
            )
            worksheet.cell(row_number, headers["reviewer_notes"]).value = (
                f"{previous_note}\n{audit_note}" if previous_note else audit_note
            )

        output_path = output_dir / source_path.name
        temporary_path = output_path.with_suffix(".tmp.xlsx")
        workbook.save(temporary_path)
        workbook.close()
        temporary_path.replace(output_path)
        locale = records[0][0]["target_locale"]
        outputs[locale] = str(output_path)
        print(f"Wrote {len(records)} translated row(s): {output_path}")

    return outputs


def collect_run(run_dir: Path) -> dict[str, Any]:
    manifest_path, manifest = load_manifest(run_dir)
    if manifest.get("batch_status") != "completed":
        manifest = status_run(run_dir, wait=False, poll_interval=30)
    if manifest.get("batch_status") != "completed":
        raise TranslationError(
            f"Batch is not complete (status={manifest.get('batch_status')!r})"
        )

    client = make_openai_client()
    output_file_id = manifest.get("output_file_id")
    if not output_file_id:
        raise TranslationError("Completed batch has no output_file_id")

    output_path = run_dir / "batch_output.jsonl"
    download_file(client, output_file_id, output_path)
    if manifest.get("error_file_id"):
        download_file(
            client,
            manifest["error_file_id"],
            run_dir / "batch_errors.jsonl",
        )

    output_rows = parse_jsonl(output_path)
    expected_ids = set(manifest["entries"])
    received_ids: set[str] = set()
    translations: dict[str, dict[str, Any]] = {}
    validation: dict[str, dict[str, Any]] = {}

    for result in output_rows:
        custom_id = result.get("custom_id")
        if not isinstance(custom_id, str):
            raise TranslationError("Batch output row is missing custom_id")
        if custom_id not in expected_ids:
            raise TranslationError(f"Unknown custom_id in output: {custom_id}")
        if custom_id in received_ids:
            raise TranslationError(f"Duplicate output custom_id: {custom_id}")
        received_ids.add(custom_id)
        entry = manifest["entries"][custom_id]

        try:
            content = extract_chat_content(result)
            translated = json.loads(content)
            translated = merge_locked_targets(
                translated,
                entry.get("locked_targets", {}),
            )
            errors = validate_translation(entry, translated)
        except (TranslationError, json.JSONDecodeError) as exc:
            translated = None
            errors = [str(exc)]

        validation[custom_id] = {
            "slug": entry["slug"],
            "target_locale": entry["target_locale"],
            "valid": not errors,
            "errors": errors,
        }
        if not errors and isinstance(translated, dict):
            translations[custom_id] = translated

    missing_ids = sorted(expected_ids - received_ids)
    for custom_id in missing_ids:
        entry = manifest["entries"][custom_id]
        validation[custom_id] = {
            "slug": entry["slug"],
            "target_locale": entry["target_locale"],
            "valid": False,
            "errors": ["no successful result line returned by Batch API"],
        }

    validation_report = {
        "run_id": manifest["run_id"],
        "batch_id": manifest["batch_id"],
        "validated_at": utc_now(),
        "expected": len(expected_ids),
        "received": len(received_ids),
        "valid": len(translations),
        "invalid": len(expected_ids) - len(translations),
        "results": validation,
    }
    atomic_write_json(run_dir / "validation_report.json", validation_report)

    if len(translations) != len(expected_ids):
        manifest["status"] = "validation_failed"
        save_manifest(manifest_path, manifest)
        raise TranslationError(
            f"Validation failed for {len(expected_ids) - len(translations)} of "
            f"{len(expected_ids)} requests; workbooks were not written"
        )

    atomic_write_json(run_dir / "validated_translations.json", translations)
    output_workbooks = write_translations(run_dir, manifest, translations)
    manifest["status"] = "collected"
    manifest["collected_at"] = utc_now()
    manifest["output_workbooks"] = output_workbooks
    save_manifest(manifest_path, manifest)
    print(f"Collected and validated {len(translations)} translations")
    return manifest


def add_prepare_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--workbook-dir",
        required=True,
        help="Directory containing medora_medical_content_{locale}.xlsx",
    )
    parser.add_argument(
        "--run-dir",
        required=True,
        help="New or resumable run directory for JSONL, manifest, and outputs",
    )
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument(
        "--languages",
        nargs="+",
        default=list(DEFAULT_LANGUAGES),
        help="Target locales, space- or comma-separated (default: ru ar id)",
    )
    parser.add_argument(
        "--slug",
        action="append",
        default=[],
        help="Only translate this slug; repeat for multiple slugs",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Maximum selected rows per language",
    )
    parser.add_argument(
        "--include-complete",
        action="store_true",
        help="Allow retranslation of rows already marked Complete",
    )
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument(
        "--force",
        action="store_true",
        help="Replace an existing unsubmitted manifest in run-dir",
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Translate Medora medical-content workbooks with the OpenAI Batch API"
        )
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    prepare_parser = subparsers.add_parser(
        "prepare",
        help="Create Batch JSONL and a resumable manifest without API calls",
    )
    add_prepare_arguments(prepare_parser)

    submit_parser = subparsers.add_parser(
        "submit",
        help="Upload the prepared JSONL and create or recover the Batch job",
    )
    submit_parser.add_argument("--run-dir", required=True)

    status_parser = subparsers.add_parser(
        "status",
        help="Refresh Batch status, optionally waiting for completion",
    )
    status_parser.add_argument("--run-dir", required=True)
    status_parser.add_argument("--wait", action="store_true")
    status_parser.add_argument("--poll-interval", type=int, default=30)

    collect_parser = subparsers.add_parser(
        "collect",
        help="Download, validate, and write translated workbook copies",
    )
    collect_parser.add_argument("--run-dir", required=True)

    run_parser = subparsers.add_parser(
        "run",
        help="Prepare, submit, wait, collect, and write workbook copies",
    )
    add_prepare_arguments(run_parser)
    run_parser.add_argument("--poll-interval", type=int, default=30)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        if args.command == "prepare":
            prepare_run(args)
        elif args.command == "submit":
            submit_run(Path(args.run_dir).expanduser().resolve())
        elif args.command == "status":
            status_run(
                Path(args.run_dir).expanduser().resolve(),
                wait=args.wait,
                poll_interval=args.poll_interval,
            )
        elif args.command == "collect":
            collect_run(Path(args.run_dir).expanduser().resolve())
        elif args.command == "run":
            prepare_run(args)
            run_dir = Path(args.run_dir).expanduser().resolve()
            submit_run(run_dir)
            status_run(
                run_dir,
                wait=True,
                poll_interval=args.poll_interval,
            )
            collect_run(run_dir)
        else:
            parser.error(f"Unknown command: {args.command}")
        return 0
    except (TranslationError, OSError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
