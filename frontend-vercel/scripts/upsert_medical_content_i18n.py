#!/usr/bin/env python3
"""Validate and idempotently upsert Medora i18n workbooks into Supabase.

The default mode is a read-only dry run. Production writes require ``--execute``.
Every execution stores the generated payload, pre-write remote rows, write
receipts, and post-write comparison report in the selected audit directory.
"""

from __future__ import annotations

import argparse
import http.client
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping


HEADER_ROW = 3
DATA_START_ROW = 4
LOCALES = ("ru", "ar", "id")
SUPPORTED_LOCALES = ("en", "es", "fr", "de", "ru", "ar", "id")
REQUIRED_PROCEDURE_TRANSLATION_FIELDS = (
    "name",
    "waiting_time",
    "cost_coverage",
    "cost_factors",
    "stay_at_hospital",
    "stay_at_hotel",
    "stay_in_china",
    "surgery_detailed_description",
    "when_is_needed",
    "preparation_before_surgery",
    "recovery_process",
    "surgery_options",
    "faqs",
    "surgery_steps",
    "recovery_steps",
)
EXPECTED_COUNTS = {
    "department_i18n": 20,
    "disease_i18n": 481,
    "procedure_i18n": 969,
}

TABLES: dict[str, dict[str, Any]] = {
    "department_i18n": {
        "sheet": "Departments",
        "parent_table": "departments",
        "parent_key": "id",
        "id_column": "department_id",
        "fields": {
            "name": "target_name",
            "name_en": "target_name_en",
            "short_desc": "target_short_desc",
            "short_desc_md": "target_short_desc_md",
        },
    },
    "disease_i18n": {
        "sheet": "Diseases",
        "parent_table": "diseases",
        "parent_key": "id",
        "id_column": "disease_id",
        "fields": {
            "name": "target_name",
            "seo_meta_title": "target_seo_meta_title",
            "seo_meta_desc": "target_seo_meta_desc",
        },
    },
    "procedure_i18n": {
        "sheet": "Procedures",
        "parent_table": "procedures",
        "parent_key": "id",
        "id_column": "procedure_id",
        "fields": {
            "name": "target_name",
            "waiting_time": "target_waiting_time",
            "cost_usd": "target_cost_usd",
            "cost_estimate": "target_cost_estimate",
            "cost_as_of": "target_cost_as_of",
            "cost_coverage": "target_cost_coverage",
            "cost_factors": "target_cost_factors",
            "stay_at_hospital": "target_stay_at_hospital",
            "stay_at_hotel": "target_stay_at_hotel",
            "stay_in_china": "target_stay_in_china",
            "surgery_detailed_description": (
                "target_surgery_detailed_description"
            ),
            "when_is_needed": "target_when_is_needed",
            "preparation_before_surgery": (
                "target_preparation_before_surgery"
            ),
            "recovery_process": "target_recovery_process",
            "surgery_options": "target_surgery_options",
            "faqs": "target_faqs_json",
            "surgery_steps": "target_surgery_steps_json",
            "recovery_steps": "target_recovery_steps_json",
        },
        "json_fields": {"faqs", "surgery_steps", "recovery_steps"},
        "boolean_fields": {"cost_estimate"},
    },
}


class UpsertError(RuntimeError):
    pass


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def json_default(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    raise TypeError(f"Cannot JSON encode {type(value).__name__}")


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(
            value,
            ensure_ascii=False,
            indent=2,
            default=json_default,
        )
        + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def normalize_cell(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def parse_boolean(value: Any, location: str) -> bool | None:
    value = normalize_cell(value)
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(value)
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"true", "yes", "1"}:
            return True
        if normalized in {"false", "no", "0"}:
            return False
    raise UpsertError(f"{location}: invalid boolean value {value!r}")


def parse_json_cell(value: Any, location: str) -> Any:
    value = normalize_cell(value)
    if value is None:
        return None
    if not isinstance(value, str):
        raise UpsertError(f"{location}: expected JSON text")
    try:
        return json.loads(value)
    except json.JSONDecodeError as exc:
        raise UpsertError(f"{location}: invalid JSON: {exc}") from exc


def is_nonempty_content(value: Any) -> bool:
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, (dict, list)):
        return bool(value)
    return value is not None


def validate_procedure_record(
    record: Mapping[str, Any],
    location: str,
) -> None:
    missing = [
        field
        for field in REQUIRED_PROCEDURE_TRANSLATION_FIELDS
        if not is_nonempty_content(record.get(field))
    ]
    if missing:
        raise UpsertError(
            f"{location}: incomplete procedure translation; missing "
            + ", ".join(missing)
        )
    cost = record.get("cost_usd")
    if not isinstance(cost, str) or not cost.strip():
        raise UpsertError(f"{location}: cost_usd must be non-empty text")
    if re.search(r"[\u3400-\u9fff]", cost):
        raise UpsertError(
            f"{location}: cost_usd contains untranslated Chinese text"
        )


def workbook_path(workbook_dir: Path, locale: str) -> Path:
    preferred = workbook_dir / f"medora_medical_content_{locale}.xlsx"
    if preferred.is_file():
        return preferred

    candidates = sorted(workbook_dir.glob(f"*_{locale}.xlsx"))
    if len(candidates) == 1:
        return candidates[0]
    if not candidates:
        raise UpsertError(f"Workbook not found: {preferred}")
    raise UpsertError(
        f"Multiple workbooks found for {locale!r}: "
        + ", ".join(str(path) for path in candidates)
    )


def worksheet_headers(worksheet: Any) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column in range(1, worksheet.max_column + 1):
        value = worksheet.cell(HEADER_ROW, column).value
        if isinstance(value, str) and value.strip():
            headers[value.strip()] = column
    return headers


def build_payloads(
    workbook_dir: Path,
    locales: Iterable[str] = LOCALES,
    tables: Iterable[str] = tuple(TABLES),
    expected_counts: Mapping[str, int] = EXPECTED_COUNTS,
) -> dict[str, dict[str, list[dict[str, Any]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise UpsertError("openpyxl is required") from exc

    selected_locales = tuple(locales)
    selected_tables = tuple(tables)
    payloads: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for locale in selected_locales:
        workbook = load_workbook(
            workbook_path(workbook_dir, locale),
            read_only=False,
            data_only=False,
        )
        locale_payload: dict[str, list[dict[str, Any]]] = {}
        for table in selected_tables:
            config = TABLES[table]
            if config["sheet"] not in workbook.sheetnames:
                raise UpsertError(
                    f"{workbook_path(workbook_dir, locale).name}: "
                    f"missing sheet {config['sheet']!r}"
                )
            worksheet = workbook[config["sheet"]]
            headers = worksheet_headers(worksheet)
            required = {
                "entity_id",
                "target_locale",
                "status",
                *config["fields"].values(),
            }
            missing = sorted(required - set(headers))
            if missing:
                raise UpsertError(
                    f"{workbook_path(workbook_dir, locale).name}:"
                    f"{config['sheet']}: missing columns {missing}"
                )

            rows: list[dict[str, Any]] = []
            seen_ids: set[str] = set()
            for row_number in range(DATA_START_ROW, worksheet.max_row + 1):
                status = str(
                    worksheet.cell(row_number, headers["status"]).value or ""
                ).strip()
                if status != "Complete":
                    continue
                target_locale = str(
                    worksheet.cell(
                        row_number,
                        headers["target_locale"],
                    ).value
                    or ""
                ).strip()
                if target_locale != locale:
                    raise UpsertError(
                        f"{config['sheet']}!row {row_number}: target locale "
                        f"{target_locale!r} does not match {locale!r}"
                    )
                entity_id = str(
                    worksheet.cell(
                        row_number,
                        headers["entity_id"],
                    ).value
                    or ""
                ).strip()
                if not entity_id:
                    raise UpsertError(
                        f"{config['sheet']}!row {row_number}: empty entity_id"
                    )
                if entity_id in seen_ids:
                    raise UpsertError(
                        f"{config['sheet']}: duplicate entity_id {entity_id}"
                    )
                seen_ids.add(entity_id)

                record: dict[str, Any] = {
                    config["id_column"]: entity_id,
                    "locale": locale,
                }
                for database_field, workbook_field in config["fields"].items():
                    value = worksheet.cell(
                        row_number,
                        headers[workbook_field],
                    ).value
                    location = (
                        f"{config['sheet']}!row {row_number} {workbook_field}"
                    )
                    if database_field in config.get("json_fields", set()):
                        record[database_field] = parse_json_cell(value, location)
                    elif database_field in config.get("boolean_fields", set()):
                        record[database_field] = parse_boolean(value, location)
                    else:
                        record[database_field] = normalize_cell(value)
                if table == "procedure_i18n":
                    validate_procedure_record(
                        record,
                        f"{config['sheet']}!row {row_number}",
                    )
                rows.append(record)

            expected = expected_counts[table]
            if len(rows) != expected:
                raise UpsertError(
                    f"{locale} {table}: expected {expected} Complete rows, "
                    f"found {len(rows)}"
                )
            locale_payload[table] = rows
        workbook.close()
        payloads[locale] = locale_payload
    return payloads


class SupabaseRest:
    def __init__(self) -> None:
        self.base_url = os.getenv("CHINA_MEDICAL_SUPABASE_URL", "").rstrip("/")
        self.service_key = os.getenv("CHINA_MEDICAL_SUPABASE_SERVICE_KEY", "")
        if not self.base_url or not self.service_key:
            raise UpsertError(
                "CHINA_MEDICAL_SUPABASE_URL and "
                "CHINA_MEDICAL_SUPABASE_SERVICE_KEY must be set"
            )

    def request(
        self,
        method: str,
        table: str,
        query: Mapping[str, str],
        body: Any = None,
        prefer: str | None = None,
    ) -> tuple[Any, Mapping[str, str]]:
        encoded_query = urllib.parse.urlencode(query, safe="(),.*")
        url = f"{self.base_url}/rest/v1/{table}?{encoded_query}"
        headers = {
            "apikey": self.service_key,
            "Authorization": f"Bearer {self.service_key}",
            "Accept": "application/json",
        }
        data = None
        if body is not None:
            data = json.dumps(
                body,
                ensure_ascii=False,
                default=json_default,
            ).encode("utf-8")
            headers["Content-Type"] = "application/json"
        if prefer:
            headers["Prefer"] = prefer
        request = urllib.request.Request(
            url,
            data=data,
            headers=headers,
            method=method,
        )
        # Supabase writes are idempotent because every POST uses the composite
        # primary key as ``on_conflict``. A short transport-only retry therefore
        # safely handles TLS EOFs without duplicating rows or altering content.
        for attempt in range(1, 4):
            try:
                with urllib.request.urlopen(request, timeout=120) as response:
                    raw = response.read()
                    parsed = json.loads(raw) if raw else None
                    return parsed, dict(response.headers.items())
            except urllib.error.HTTPError as exc:
                details = exc.read().decode("utf-8", errors="replace")
                transient = exc.code in {408, 425, 429, 500, 502, 503, 504}
                if transient and attempt < 3:
                    time.sleep(2 ** (attempt - 1))
                    continue
                raise UpsertError(
                    f"{method} {table} failed with HTTP {exc.code}: {details}"
                ) from exc
            except (
                urllib.error.URLError,
                http.client.IncompleteRead,
                http.client.RemoteDisconnected,
                ConnectionResetError,
                TimeoutError,
            ) as exc:
                if attempt < 3:
                    time.sleep(2 ** (attempt - 1))
                    continue
                raise UpsertError(
                    f"{method} {table} failed after 3 transport attempts: {exc}"
                ) from exc
        raise AssertionError("unreachable")

    def select_locale(
        self,
        table: str,
        locale: str,
        fields: Iterable[str],
    ) -> list[dict[str, Any]]:
        selected_fields = list(fields)
        page_size = 100
        all_rows: list[dict[str, Any]] = []
        offset = 0
        while True:
            rows, _ = self.request(
                "GET",
                table,
                {
                    "select": ",".join(selected_fields),
                    "locale": f"eq.{locale}",
                    "order": f"{selected_fields[0]}.asc",
                    "limit": str(page_size),
                    "offset": str(offset),
                },
            )
            if not isinstance(rows, list):
                raise UpsertError(f"{table}: expected an array response")
            all_rows.extend(rows)
            if len(rows) < page_size:
                return all_rows
            offset += page_size

    def select_target_rows(
        self,
        table: str,
        locale: str,
        id_column: str,
        ids: Iterable[str],
        fields: Iterable[str],
    ) -> list[dict[str, Any]]:
        """Read only explicitly targeted composite keys for verification."""
        selected_fields = list(fields)
        wanted_ids = sorted(set(ids))
        all_rows: list[dict[str, Any]] = []
        for id_chunk in chunks(wanted_ids, 100):
            rows, _ = self.request(
                "GET",
                table,
                {
                    "select": ",".join(selected_fields),
                    "locale": f"eq.{locale}",
                    id_column: f"in.({','.join(id_chunk)})",
                    "order": f"{id_column}.asc",
                    "limit": str(len(id_chunk)),
                },
            )
            if not isinstance(rows, list):
                raise UpsertError(f"{table}: expected an array response")
            all_rows.extend(rows)
        # A duplicate composite key indicates corrupt or unexpectedly broad
        # read-back and must fail before any comparison is trusted.
        index_rows(all_rows, id_column)
        return all_rows

    def select_locale_count(
        self,
        table: str,
        locale: str,
        id_column: str,
    ) -> int:
        """Return the total remote row count for a locale without comparing it."""
        return len(self.select_locale(table, locale, (id_column, "locale")))

    def select_parent_ids(
        self,
        table: str,
        key: str,
    ) -> set[str]:
        rows, _ = self.request(
            "GET",
            table,
            {"select": key, "limit": "5000"},
        )
        if not isinstance(rows, list):
            raise UpsertError(f"{table}: expected an array response")
        return {str(row[key]) for row in rows}

    def upsert(
        self,
        table: str,
        id_column: str,
        rows: list[dict[str, Any]],
    ) -> None:
        self.request(
            "POST",
            table,
            {"on_conflict": f"{id_column},locale"},
            body=rows,
            prefer="resolution=merge-duplicates,return=minimal",
        )


def normalized_record(record: Mapping[str, Any]) -> dict[str, Any]:
    return json.loads(
        json.dumps(
            record,
            ensure_ascii=False,
            sort_keys=True,
            default=json_default,
        )
    )


def index_rows(rows: Iterable[Mapping[str, Any]], id_column: str) -> dict[str, dict[str, Any]]:
    indexed: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = f"{row[id_column]}|{row['locale']}"
        if key in indexed:
            raise UpsertError(f"Duplicate remote row {key}")
        indexed[key] = normalized_record(row)
    return indexed


def compare_rows(
    expected: list[dict[str, Any]],
    actual: list[dict[str, Any]],
    id_column: str,
) -> dict[str, Any]:
    expected_index = index_rows(expected, id_column)
    actual_index = index_rows(actual, id_column)
    missing = sorted(set(expected_index) - set(actual_index))
    extra = sorted(set(actual_index) - set(expected_index))
    mismatched: list[dict[str, Any]] = []
    for key in sorted(set(expected_index) & set(actual_index)):
        if expected_index[key] != actual_index[key]:
            mismatched.append(
                {
                    "key": key,
                    "expected": expected_index[key],
                    "actual": actual_index[key],
                }
            )
    return {
        "expected_count": len(expected_index),
        "actual_count": len(actual_index),
        "missing_keys": missing,
        "extra_keys": extra,
        "mismatched": mismatched,
        "matches": not missing and not extra and not mismatched,
    }


def rows_requiring_upsert(
    expected: list[dict[str, Any]],
    actual: list[dict[str, Any]],
    id_column: str,
) -> list[dict[str, Any]]:
    """Return only absent or non-identical rows for safe resumable execution."""
    actual_index = index_rows(actual, id_column)
    required: list[dict[str, Any]] = []
    for row in expected:
        key = f"{row[id_column]}|{row['locale']}"
        if actual_index.get(key) != normalized_record(row):
            required.append(row)
    return required


def chunks(values: list[Any], size: int) -> Iterable[list[Any]]:
    for start in range(0, len(values), size):
        yield values[start : start + size]


def create_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook-dir", required=True)
    parser.add_argument("--audit-dir", required=True)
    parser.add_argument(
        "--execute",
        action="store_true",
        help="Perform production upserts; omitted means read-only dry run",
    )
    parser.add_argument("--chunk-size", type=int, default=100)
    parser.add_argument(
        "--locales",
        nargs="+",
        choices=SUPPORTED_LOCALES,
        help=(
            "Locales to process. Omit to preserve the existing default: "
            "ru ar id."
        ),
    )
    parser.add_argument(
        "--tables",
        nargs="+",
        choices=tuple(TABLES),
        help=(
            "i18n tables to process. Omit to preserve the existing default "
            "of all three tables."
        ),
    )
    parser.add_argument(
        "--expected-procedure-count",
        type=int,
        help=(
            "Exact Complete-row count required per selected Procedures "
            "workbook. This override is allowed only with "
            "--tables procedure_i18n."
        ),
    )
    return parser


def resolve_selection(
    args: argparse.Namespace,
) -> tuple[tuple[str, ...], tuple[str, ...], dict[str, int], bool]:
    locales = tuple(args.locales) if args.locales else LOCALES
    tables = tuple(args.tables) if args.tables else tuple(TABLES)
    if len(set(locales)) != len(locales):
        raise UpsertError("--locales contains duplicate values")
    if len(set(tables)) != len(tables):
        raise UpsertError("--tables contains duplicate values")

    expected_counts = dict(EXPECTED_COUNTS)
    if args.expected_procedure_count is not None:
        if args.expected_procedure_count < 1:
            raise UpsertError("--expected-procedure-count must be positive")
        if tables != ("procedure_i18n",):
            raise UpsertError(
                "--expected-procedure-count requires exactly "
                "--tables procedure_i18n"
            )
        expected_counts["procedure_i18n"] = args.expected_procedure_count

    targeted = any(
        (
            args.locales is not None,
            args.tables is not None,
            args.expected_procedure_count is not None,
        )
    )
    return locales, tables, expected_counts, targeted


def validate_payload_composite_keys(
    payloads: Mapping[str, Mapping[str, list[dict[str, Any]]]],
    locales: Iterable[str],
    tables: Iterable[str],
) -> None:
    for table in tables:
        id_column = TABLES[table]["id_column"]
        seen: set[str] = set()
        for locale in locales:
            for row in payloads[locale][table]:
                key = f"{row[id_column]}|{row['locale']}"
                if key in seen:
                    raise UpsertError(f"Duplicate payload composite key {key}")
                seen.add(key)


def target_remote_rows(
    client: Any,
    table: str,
    locale: str,
    expected: list[dict[str, Any]],
    targeted: bool,
) -> list[dict[str, Any]]:
    config = TABLES[table]
    fields = list(expected[0])
    if not targeted:
        return client.select_locale(table, locale, fields)
    return client.select_target_rows(
        table,
        locale,
        config["id_column"],
        (row[config["id_column"]] for row in expected),
        fields,
    )


def ensure_upserts_are_targeted(
    rows: list[dict[str, Any]],
    expected: list[dict[str, Any]],
    id_column: str,
) -> None:
    allowed = set(index_rows(expected, id_column))
    attempted = set(index_rows(rows, id_column))
    outside_target = sorted(attempted - allowed)
    if outside_target:
        raise UpsertError(
            f"Refusing to upsert non-target composite keys: "
            f"{outside_target[:10]}"
        )


def run(args: argparse.Namespace, client: Any | None = None) -> int:
    locales, tables, expected_counts, targeted = resolve_selection(args)

    if args.chunk_size < 1 or args.chunk_size > 500:
        raise UpsertError("--chunk-size must be between 1 and 500")

    workbook_dir = Path(args.workbook_dir).expanduser().resolve()
    audit_dir = Path(args.audit_dir).expanduser().resolve()
    audit_dir.mkdir(parents=True, exist_ok=True)
    payloads = build_payloads(
        workbook_dir,
        locales=locales,
        tables=tables,
        expected_counts=expected_counts,
    )
    validate_payload_composite_keys(payloads, locales, tables)
    generated = {
        "version": 2,
        "generated_at": utc_now(),
        "workbook_dir": str(workbook_dir),
        "targeted": targeted,
        "locales": locales,
        "tables": tables,
        "expected_counts": {
            table: expected_counts[table] for table in tables
        },
        "payloads": payloads,
    }
    write_json(audit_dir / "upsert_payloads.json", generated)

    if client is None:
        client = SupabaseRest()
    parent_ids: dict[str, set[str]] = {}
    for table in tables:
        config = TABLES[table]
        parent_ids[table] = client.select_parent_ids(
            config["parent_table"],
            config["parent_key"],
        )
        wanted = {
            row[config["id_column"]]
            for locale in locales
            for row in payloads[locale][table]
        }
        missing = sorted(wanted - parent_ids[table])
        if missing:
            raise UpsertError(
                f"{table}: {len(missing)} IDs are absent from "
                f"{config['parent_table']}: {missing[:10]}"
            )

    before: dict[str, dict[str, list[dict[str, Any]]]] = {}
    remote_total_before_counts: dict[str, dict[str, int]] = {}
    expected_remote_after_counts: dict[str, dict[str, int]] = {}
    for locale in locales:
        before[locale] = {}
        remote_total_before_counts[locale] = {}
        expected_remote_after_counts[locale] = {}
        for table in tables:
            config = TABLES[table]
            expected = payloads[locale][table]
            before[locale][table] = target_remote_rows(
                client,
                table,
                locale,
                expected,
                targeted,
            )
            if targeted:
                remote_total_before_counts[locale][table] = (
                    client.select_locale_count(
                        table,
                        locale,
                        config["id_column"],
                    )
                )
            else:
                remote_total_before_counts[locale][table] = len(
                    before[locale][table]
                )
            expected_index = index_rows(expected, config["id_column"])
            before_index = index_rows(
                before[locale][table],
                config["id_column"],
            )
            missing_target_count = len(set(expected_index) - set(before_index))
            expected_remote_after_counts[locale][table] = (
                remote_total_before_counts[locale][table]
                + missing_target_count
            )
    write_json(
        audit_dir / "remote_before.json",
        {
            "captured_at": utc_now(),
            "target_rows": before,
            "remote_total_locale_counts": remote_total_before_counts,
        },
    )

    plan = {
        "mode": "execute" if args.execute else "dry-run",
        "targeted": targeted,
        "locales": locales,
        "tables": tables,
        "generated_at": utc_now(),
        "parent_id_counts": {
            table: len(values) for table, values in parent_ids.items()
        },
        "planned_counts": {
            locale: {
                table: len(rows)
                for table, rows in locale_payload.items()
            }
            for locale, locale_payload in payloads.items()
        },
        "target_rows_found_before": {
            locale: {table: len(rows) for table, rows in locale_rows.items()}
            for locale, locale_rows in before.items()
        },
        "remote_total_locale_counts_before": remote_total_before_counts,
        "expected_remote_total_locale_counts_after": (
            expected_remote_after_counts
        ),
    }
    write_json(audit_dir / "plan.json", plan)
    print(json.dumps(plan, ensure_ascii=False, indent=2))
    if not args.execute:
        print("Dry run complete; no database rows were changed.")
        return 0

    receipts: list[dict[str, Any]] = []
    for locale in locales:
        for table in tables:
            config = TABLES[table]
            table_rows = rows_requiring_upsert(
                payloads[locale][table],
                before[locale][table],
                config["id_column"],
            )
            ensure_upserts_are_targeted(
                table_rows,
                payloads[locale][table],
                config["id_column"],
            )
            print(
                f"{locale} {table}: {len(table_rows)} row(s) require upsert; "
                f"{len(payloads[locale][table]) - len(table_rows)} already match"
            )
            for chunk_number, chunk in enumerate(
                chunks(table_rows, args.chunk_size),
                start=1,
            ):
                client.upsert(
                    table,
                    config["id_column"],
                    chunk,
                )
                receipt = {
                    "at": utc_now(),
                    "locale": locale,
                    "table": table,
                    "chunk": chunk_number,
                    "rows": len(chunk),
                }
                receipts.append(receipt)
                write_json(audit_dir / "write_receipts.json", receipts)
                print(
                    f"Upserted {locale} {table} chunk {chunk_number}: "
                    f"{len(chunk)} row(s)"
                )

    comparisons: dict[str, dict[str, Any]] = {}
    remote_after: dict[str, dict[str, list[dict[str, Any]]]] = {}
    remote_total_after_counts: dict[str, dict[str, int]] = {}
    remote_total_count_checks: dict[str, dict[str, dict[str, Any]]] = {}
    for locale in locales:
        comparisons[locale] = {}
        remote_after[locale] = {}
        remote_total_after_counts[locale] = {}
        remote_total_count_checks[locale] = {}
        for table in tables:
            config = TABLES[table]
            expected = payloads[locale][table]
            actual = target_remote_rows(
                client,
                table,
                locale,
                expected,
                targeted,
            )
            remote_after[locale][table] = actual
            comparisons[locale][table] = compare_rows(
                expected,
                actual,
                config["id_column"],
            )
            if targeted:
                actual_total_count = client.select_locale_count(
                    table,
                    locale,
                    config["id_column"],
                )
            else:
                actual_total_count = len(actual)
            remote_total_after_counts[locale][table] = actual_total_count
            expected_total_count = expected_remote_after_counts[locale][table]
            remote_total_count_checks[locale][table] = {
                "before": remote_total_before_counts[locale][table],
                "expected_after": expected_total_count,
                "actual_after": actual_total_count,
                "matches": actual_total_count == expected_total_count,
            }
    write_json(
        audit_dir / "remote_after.json",
        {
            "captured_at": utc_now(),
            "target_rows": remote_after,
            "remote_total_locale_counts": remote_total_after_counts,
        },
    )
    target_rows_match = all(
        result["matches"]
        for locale_result in comparisons.values()
        for result in locale_result.values()
    )
    remote_total_counts_match = all(
        result["matches"]
        for locale_result in remote_total_count_checks.values()
        for result in locale_result.values()
    )
    report = {
        "verified_at": utc_now(),
        "targeted": targeted,
        "comparisons": comparisons,
        "remote_total_count_checks": remote_total_count_checks,
        "target_rows_match": target_rows_match,
        "remote_total_counts_match": remote_total_counts_match,
        "all_match": target_rows_match and remote_total_counts_match,
    }
    write_json(audit_dir / "verification.json", report)
    if not report["all_match"]:
        raise UpsertError(
            "Post-write verification failed; inspect verification.json"
        )
    print("Production upsert and exact read-back verification succeeded.")
    return 0


def main() -> int:
    return run(create_parser().parse_args())


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except UpsertError as exc:
        print(f"error: {exc}", file=sys.stderr)
        raise SystemExit(1)
