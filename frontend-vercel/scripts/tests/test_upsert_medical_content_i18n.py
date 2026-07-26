from __future__ import annotations

import argparse
import copy
import importlib.util
import json
import tempfile
import unittest
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook


SCRIPT_PATH = (
    Path(__file__).resolve().parents[1] / "upsert_medical_content_i18n.py"
)
SPEC = importlib.util.spec_from_file_location(
    "upsert_medical_content_i18n",
    SCRIPT_PATH,
)
assert SPEC and SPEC.loader
upsert_module = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(upsert_module)


class FakeSupabaseRest:
    def __init__(
        self,
        parent_ids: Iterable[str],
        remote_rows: Iterable[dict[str, Any]],
    ) -> None:
        self.parent_ids = set(parent_ids)
        self.rows = {
            "procedure_i18n": [copy.deepcopy(row) for row in remote_rows]
        }
        self.upsert_calls: list[dict[str, Any]] = []

    @staticmethod
    def _project(
        rows: Iterable[dict[str, Any]],
        fields: Iterable[str],
    ) -> list[dict[str, Any]]:
        selected = tuple(fields)
        return [
            {field: copy.deepcopy(row.get(field)) for field in selected}
            for row in rows
        ]

    def select_parent_ids(self, table: str, key: str) -> set[str]:
        self.last_parent_request = (table, key)
        return set(self.parent_ids)

    def select_locale(
        self,
        table: str,
        locale: str,
        fields: Iterable[str],
    ) -> list[dict[str, Any]]:
        return self._project(
            (
                row
                for row in self.rows[table]
                if row["locale"] == locale
            ),
            fields,
        )

    def select_target_rows(
        self,
        table: str,
        locale: str,
        id_column: str,
        ids: Iterable[str],
        fields: Iterable[str],
    ) -> list[dict[str, Any]]:
        wanted = set(ids)
        return self._project(
            (
                row
                for row in self.rows[table]
                if row["locale"] == locale
                and row[id_column] in wanted
            ),
            fields,
        )

    def select_locale_count(
        self,
        table: str,
        locale: str,
        id_column: str,
    ) -> int:
        del id_column
        return sum(
            row["locale"] == locale for row in self.rows[table]
        )

    def upsert(
        self,
        table: str,
        id_column: str,
        rows: list[dict[str, Any]],
    ) -> None:
        self.upsert_calls.append(
            {
                "table": table,
                "id_column": id_column,
                "rows": copy.deepcopy(rows),
            }
        )
        indexed = {
            (row[id_column], row["locale"]): index
            for index, row in enumerate(self.rows[table])
        }
        for row in rows:
            key = (row[id_column], row["locale"])
            if key in indexed:
                self.rows[table][indexed[key]] = copy.deepcopy(row)
            else:
                indexed[key] = len(self.rows[table])
                self.rows[table].append(copy.deepcopy(row))


def procedure_headers() -> list[str]:
    config = upsert_module.TABLES["procedure_i18n"]
    return [
        "entity_id",
        "target_locale",
        "status",
        *config["fields"].values(),
    ]


def target_values(seed: str) -> dict[str, Any]:
    config = upsert_module.TABLES["procedure_i18n"]
    values: dict[str, Any] = {}
    for database_field, workbook_field in config["fields"].items():
        if database_field in config["json_fields"]:
            values[workbook_field] = json.dumps(
                {"seed": seed},
                ensure_ascii=False,
            )
        elif database_field in config["boolean_fields"]:
            values[workbook_field] = True
        else:
            values[workbook_field] = f"{workbook_field}-{seed}"
    return values


def expected_record(entity_id: str, locale: str) -> dict[str, Any]:
    config = upsert_module.TABLES["procedure_i18n"]
    source = target_values(f"{entity_id}-{locale}")
    record = {
        "procedure_id": entity_id,
        "locale": locale,
    }
    for database_field, workbook_field in config["fields"].items():
        value = source[workbook_field]
        if database_field in config["json_fields"]:
            value = json.loads(value)
        record[database_field] = value
    return record


def write_procedure_workbook(
    directory: Path,
    locale: str,
    rows: list[tuple[str, str]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Procedures"
    headers = procedure_headers()
    for column, header in enumerate(headers, start=1):
        worksheet.cell(3, column).value = header
    for row_number, (entity_id, status) in enumerate(rows, start=4):
        values = {
            "entity_id": entity_id,
            "target_locale": locale,
            "status": status,
            **target_values(f"{entity_id}-{locale}"),
        }
        for column, header in enumerate(headers, start=1):
            worksheet.cell(row_number, column).value = values[header]
    workbook.save(directory / f"medora_medical_content_{locale}.xlsx")


class SelectionTests(unittest.TestCase):
    def test_workbook_resolution_prefers_legacy_name_then_unique_backfill(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            backfill = root / "medora_procedure_backfill_en.xlsx"
            backfill.touch()
            self.assertEqual(
                upsert_module.workbook_path(root, "en"),
                backfill,
            )

            preferred = root / "medora_medical_content_en.xlsx"
            preferred.touch()
            self.assertEqual(
                upsert_module.workbook_path(root, "en"),
                preferred,
            )

    def test_workbook_resolution_rejects_ambiguous_fallbacks(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            (root / "first_en.xlsx").touch()
            (root / "second_en.xlsx").touch()
            with self.assertRaisesRegex(
                upsert_module.UpsertError,
                "Multiple workbooks found",
            ):
                upsert_module.workbook_path(root, "en")

    def test_no_new_flags_preserves_existing_defaults(self) -> None:
        args = upsert_module.create_parser().parse_args(
            ["--workbook-dir", "/tmp/workbooks", "--audit-dir", "/tmp/audit"]
        )
        locales, tables, counts, targeted = (
            upsert_module.resolve_selection(args)
        )
        self.assertEqual(locales, ("ru", "ar", "id"))
        self.assertEqual(tables, tuple(upsert_module.TABLES))
        self.assertEqual(counts, upsert_module.EXPECTED_COUNTS)
        self.assertFalse(targeted)
        self.assertFalse(args.execute)

    def test_expected_procedure_override_requires_procedure_only(self) -> None:
        args = argparse.Namespace(
            locales=["en"],
            tables=["procedure_i18n", "disease_i18n"],
            expected_procedure_count=71,
        )
        with self.assertRaisesRegex(
            upsert_module.UpsertError,
            "requires exactly --tables procedure_i18n",
        ):
            upsert_module.resolve_selection(args)


class ProcedureOnlyBackfillTests(unittest.TestCase):
    def test_incomplete_procedure_rows_fail_before_any_write(self) -> None:
        cases = (
            ("target_recovery_process", None, "missing recovery_process"),
            ("target_faqs_json", "{}", "missing faqs"),
            (
                "target_cost_usd",
                "约 ¥36,400（估算）",
                "contains untranslated Chinese text",
            ),
        )
        for workbook_field, invalid_value, expected_error in cases:
            with self.subTest(workbook_field=workbook_field):
                with tempfile.TemporaryDirectory() as temporary:
                    root = Path(temporary)
                    workbook_dir = root / "workbooks"
                    workbook_dir.mkdir()
                    write_procedure_workbook(
                        workbook_dir,
                        "en",
                        [("procedure-a", "Complete")],
                    )
                    workbook_path = (
                        workbook_dir / "medora_medical_content_en.xlsx"
                    )
                    workbook = load_workbook(workbook_path)
                    worksheet = workbook["Procedures"]
                    headers = upsert_module.worksheet_headers(worksheet)
                    worksheet.cell(
                        4,
                        headers[workbook_field],
                    ).value = invalid_value
                    workbook.save(workbook_path)
                    workbook.close()

                    fake = FakeSupabaseRest(("procedure-a",), ())
                    args = argparse.Namespace(
                        workbook_dir=str(workbook_dir),
                        audit_dir=str(root / "audit"),
                        execute=True,
                        chunk_size=100,
                        locales=["en"],
                        tables=["procedure_i18n"],
                        expected_procedure_count=1,
                    )
                    with self.assertRaisesRegex(
                        upsert_module.UpsertError,
                        expected_error,
                    ):
                        upsert_module.run(args, client=fake)
                    self.assertEqual(fake.upsert_calls, [])

    def test_execute_only_writes_target_keys_and_exactly_verifies_them(
        self,
    ) -> None:
        locales = upsert_module.SUPPORTED_LOCALES
        target_ids = ("procedure-a", "procedure-b")
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            workbook_dir = root / "workbooks"
            audit_dir = root / "audit"
            workbook_dir.mkdir()
            for locale in locales:
                write_procedure_workbook(
                    workbook_dir,
                    locale,
                    [
                        (target_ids[0], "Complete"),
                        (target_ids[1], "Complete"),
                        (f"ignored-{locale}", "Not Started"),
                    ],
                )

            remote_rows: list[dict[str, Any]] = []
            for locale in locales:
                stale = expected_record(target_ids[0], locale)
                stale["name"] = "stale remote value"
                remote_rows.extend(
                    [
                        stale,
                        {
                            "procedure_id": f"outside-{locale}",
                            "locale": locale,
                            "name": f"untouched-{locale}",
                        },
                    ]
                )
            fake = FakeSupabaseRest(target_ids, remote_rows)
            non_target_before = {
                (row["procedure_id"], row["locale"]): copy.deepcopy(row)
                for row in fake.rows["procedure_i18n"]
                if row["procedure_id"].startswith("outside-")
            }
            args = argparse.Namespace(
                workbook_dir=str(workbook_dir),
                audit_dir=str(audit_dir),
                execute=True,
                chunk_size=3,
                locales=list(locales),
                tables=["procedure_i18n"],
                expected_procedure_count=2,
            )

            self.assertEqual(upsert_module.run(args, client=fake), 0)

            written = [
                row
                for call in fake.upsert_calls
                for row in call["rows"]
            ]
            self.assertEqual(len(written), len(locales) * len(target_ids))
            self.assertEqual(
                {
                    (row["procedure_id"], row["locale"])
                    for row in written
                },
                {
                    (entity_id, locale)
                    for locale in locales
                    for entity_id in target_ids
                },
            )
            non_target_after = {
                (row["procedure_id"], row["locale"]): row
                for row in fake.rows["procedure_i18n"]
                if row["procedure_id"].startswith("outside-")
            }
            self.assertEqual(non_target_after, non_target_before)

            verification = json.loads(
                (audit_dir / "verification.json").read_text("utf-8")
            )
            self.assertTrue(verification["targeted"])
            self.assertTrue(verification["target_rows_match"])
            self.assertTrue(verification["remote_total_counts_match"])
            self.assertTrue(verification["all_match"])
            for locale in locales:
                count_check = verification["remote_total_count_checks"][
                    locale
                ]["procedure_i18n"]
                self.assertEqual(count_check["before"], 2)
                self.assertEqual(count_check["expected_after"], 3)
                self.assertEqual(count_check["actual_after"], 3)

    def test_dry_run_never_calls_upsert(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            workbook_dir = root / "workbooks"
            workbook_dir.mkdir()
            write_procedure_workbook(
                workbook_dir,
                "ar",
                [("procedure-a", "Complete")],
            )
            fake = FakeSupabaseRest(("procedure-a",), ())
            args = argparse.Namespace(
                workbook_dir=str(workbook_dir),
                audit_dir=str(root / "audit"),
                execute=False,
                chunk_size=100,
                locales=["ar"],
                tables=["procedure_i18n"],
                expected_procedure_count=1,
            )

            self.assertEqual(upsert_module.run(args, client=fake), 0)
            self.assertEqual(fake.upsert_calls, [])

    def test_duplicate_complete_parent_id_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            workbook_dir = Path(temporary)
            write_procedure_workbook(
                workbook_dir,
                "ar",
                [
                    ("procedure-a", "Complete"),
                    ("procedure-a", "Complete"),
                ],
            )
            with self.assertRaisesRegex(
                upsert_module.UpsertError,
                "duplicate entity_id procedure-a",
            ):
                upsert_module.build_payloads(
                    workbook_dir,
                    locales=("ar",),
                    tables=("procedure_i18n",),
                    expected_counts={"procedure_i18n": 2},
                )

    def test_missing_parent_id_stops_before_any_write(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            workbook_dir = root / "workbooks"
            workbook_dir.mkdir()
            write_procedure_workbook(
                workbook_dir,
                "ar",
                [("procedure-a", "Complete")],
            )
            fake = FakeSupabaseRest((), ())
            args = argparse.Namespace(
                workbook_dir=str(workbook_dir),
                audit_dir=str(root / "audit"),
                execute=True,
                chunk_size=100,
                locales=["ar"],
                tables=["procedure_i18n"],
                expected_procedure_count=1,
            )

            with self.assertRaisesRegex(
                upsert_module.UpsertError,
                "absent from procedures",
            ):
                upsert_module.run(args, client=fake)
            self.assertEqual(fake.upsert_calls, [])


if __name__ == "__main__":
    unittest.main()
